"""
app/extraction/xlsx_extractor.py

Multimodal XLSX extraction: each sheet becomes a heading + one or more
TABLE blocks (large sheets are split into row-chunks so a single block
doesn't blow past the embedding model's context), plus any embedded images
with OCR. Sheet name is preserved in metadata so retrieval can filter by
sheet.

Does not touch app/ingestion/pipeline.py.
"""

from __future__ import annotations

import os
import uuid

import openpyxl
from PIL import Image
import pytesseract

from app.extraction.base import (
    BaseExtractor,
    Block,
    BlockType,
    ExtractionResult,
)

MAX_ROWS_PER_TABLE_BLOCK = 40  # keep individual table blocks embeddable


class XlsxExtractor(BaseExtractor):
    file_types = ("xlsx", "xlsm")

    def __init__(self, image_output_dir: str = "extracted_images"):
        self.image_output_dir = image_output_dir

    def extract(self, file_path: str, doc_id: str) -> ExtractionResult:
        result = ExtractionResult(
            doc_id=doc_id,
            source_file=os.path.basename(file_path),
            file_type="xlsx",
        )

        doc_image_dir = os.path.join(self.image_output_dir, doc_id)
        os.makedirs(doc_image_dir, exist_ok=True)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            result.warnings.append(f"Failed to open XLSX: {e}")
            return result

        order_index = 0
        image_counter = 0

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            heading_block = Block(
                block_id=str(uuid.uuid4()),
                doc_id=doc_id,
                source_file=result.source_file,
                block_type=BlockType.HEADING,
                content=f"Sheet: {sheet_name}",
                order_index=order_index,
                extra_metadata={"sheet_name": sheet_name},
            )
            result.blocks.append(heading_block)
            order_index += 1

            rows = self._read_rows(sheet)
            table_blocks, order_index = self._chunk_rows_to_blocks(
                rows, doc_id, result.source_file, sheet_name, order_index, heading_block.block_id
            )
            result.blocks.extend(table_blocks)

            # Embedded images (charts render as images in some openpyxl
            # versions; native chart objects are skipped here since they
            # are not a bitmap -- flagged as a known limitation).
            try:
                for img in getattr(sheet, "_images", []):
                    image_counter += 1
                    image_filename = f"{sheet_name}_img{image_counter}.png"
                    image_path = os.path.join(doc_image_dir, image_filename)
                    with open(image_path, "wb") as f:
                        f.write(img._data())

                    ocr_text = self._run_ocr(image_path, result.warnings)

                    image_block = Block(
                        block_id=str(uuid.uuid4()),
                        doc_id=doc_id,
                        source_file=result.source_file,
                        block_type=BlockType.IMAGE,
                        content="",
                        image_path=image_path,
                        order_index=order_index,
                        extra_metadata={"sheet_name": sheet_name},
                        relates_to=[heading_block.block_id],
                    )
                    heading_block.relates_to.append(image_block.block_id)
                    result.blocks.append(image_block)
                    order_index += 1

                    if ocr_text.strip():
                        ocr_block = Block(
                            block_id=str(uuid.uuid4()),
                            doc_id=doc_id,
                            source_file=result.source_file,
                            block_type=BlockType.OCR_TEXT,
                            content=ocr_text,
                            order_index=order_index,
                            extra_metadata={"sheet_name": sheet_name},
                            relates_to=[image_block.block_id],
                        )
                        image_block.relates_to.append(ocr_block.block_id)
                        result.blocks.append(ocr_block)
                        order_index += 1
            except Exception as e:
                result.warnings.append(f"Image extraction failed on sheet {sheet_name}: {e}")

        return result

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _read_rows(sheet) -> list[list[str]]:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if row is None or all(cell is None for cell in row):
                continue
            rows.append(["" if cell is None else str(cell) for cell in row])
        return rows

    def _chunk_rows_to_blocks(
        self, rows, doc_id, source_file, sheet_name, order_index, heading_id
    ):
        blocks = []
        if not rows:
            return blocks, order_index

        header = rows[0]
        body = rows[1:]

        for start in range(0, max(len(body), 1), MAX_ROWS_PER_TABLE_BLOCK):
            chunk = body[start:start + MAX_ROWS_PER_TABLE_BLOCK]
            if not chunk and start > 0:
                break
            markdown = self._rows_to_markdown(header, chunk)

            block = Block(
                block_id=str(uuid.uuid4()),
                doc_id=doc_id,
                source_file=source_file,
                block_type=BlockType.TABLE,
                content=markdown,
                structured_data=[header] + chunk,
                order_index=order_index,
                extra_metadata={
                    "sheet_name": sheet_name,
                    "row_range": [start + 2, start + 1 + len(chunk)],  # 1-indexed, +header row
                },
                relates_to=[heading_id],
            )
            blocks.append(block)
            order_index += 1

        return blocks, order_index

    @staticmethod
    def _rows_to_markdown(header: list[str], body: list[list[str]]) -> str:
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * len(header)) + " |",
        ]
        for row in body:
            lines.append("| " + " | ".join(row) + " |")
        return "\n".join(lines)

    @staticmethod
    def _run_ocr(image_path: str, warnings: list) -> str:
        try:
            with Image.open(image_path) as im:
                return pytesseract.image_to_string(im)
        except Exception as e:
            warnings.append(f"OCR failed for {image_path}: {e}")
            return ""